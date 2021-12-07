from openpyxl import load_workbook
from openpyxl.styles import  PatternFill,Alignment,Border,Side

def excelFormat():
    excelFile='d:\\Issues\\LBY_Issue_List_LiHanmin_2021.xlsx'
    wb = load_workbook(excelFile)
    sht = wb.get_sheet_by_name('Issues')
    #设置所有行高为30 
    for i in range(sht.min_row+1,sht.max_row+1):
        sht.row_dimensions[i].height = 30
    # 首行高度
    sht.row_dimensions[1].height = 20

    # 设置列宽
    sht.column_dimensions['A'].width=4
    sht.column_dimensions['B'].width=12  # 时间列
    sht.column_dimensions['D'].width=40
    sht.column_dimensions['G'].width=55

    border = Border(left=Side(border_style='thin',color='000000'),\
        right=Side(border_style='thin',color='000000'),\
            top=Side(border_style='thin',color='000000'),
            bottom=Side(border_style='thin',color='000000'))

    # 设置格式 
    for r in sht:
        for c in r:
            if c.row != 1 :
                if(c.col_idx == 2 ):
                    c.number_format = 'yyyy/mm/dd' # 时间列修改格式
                    c.alignment = Alignment(horizontal='center', vertical='center')
                if(c.col_idx == 1 or c.col_idx == 3  or c.col_idx == 5  or c.col_idx == 6  or c.col_idx == 8   ):
                    c.alignment = Alignment(horizontal='center', vertical='center')
                if(c.col_idx == 4 or c.col_idx == 7 ):
                    c.alignment = Alignment(horizontal='left', vertical='top',wrapText=True)
                c.border = border

    # Status列颜色填充
    # light green, red, yellow 
    Color=['c6efce','ffc7ce','ffeb9c',]   #绿
    green_fill = PatternFill("solid", fgColor=Color[0])
    red_fill = PatternFill("solid", fgColor=Color[1])
    yellow_fill = PatternFill("solid", fgColor=Color[2])

    for r in sht:
        for c in r:
            # 第6列是STATUS，根据值设置填充颜色
            if c.col_idx == 6 and  c.row != 1 :
                if c.value == "Open":
                    c.fill = red_fill
                elif c.value == "Pending":
                    c.fill = yellow_fill
                elif c.value == "Closed":
                    c.fill =  green_fill 

    #  列添加筛选 
    sht.auto_filter.ref = sht.dimensions

    try:
        wb.save(excelFile) 
    except Exception as e:
        print(e)
